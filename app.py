from flask import Flask, render_template, request, redirect, session
from flask_mysqldb import MySQL
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from flask import send_file
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
import re
import os
from datetime import date, datetime

load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))


app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

app.config['MYSQL_HOST'] = os.getenv('MYSQL_HOST')
app.config['MYSQL_USER'] = os.getenv('MYSQL_USER')
app.config['MYSQL_PASSWORD'] = os.getenv('MYSQL_PASSWORD')
app.config['MYSQL_DB'] = os.getenv('MYSQL_DB')

mysql = MySQL(app)

@app.route('/')
def inicio():
    if 'usuario' in session:
        return redirect('/dashboard')
    return redirect('/login')

@app.route('/empleados')
def empleados():
    if 'usuario' not in session:
         return redirect('/login')   #listo la redireccion a login
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM empleados")
    empleados = cur.fetchall()
    cur.close()
    return render_template('empleados.html', empleados=empleados)

@app.route('/agregar_empleado', methods=['GET', 'POST'])
def agregar_empleado():
    if 'usuario' not in session:
        return redirect('/login') #listo la redireccion a login
    if request.method == 'POST':
        nombre = request.form['nombre']
        cedula = request.form['cedula']
        cargo = request.form['cargo']
        valor_dia = request.form['valor_dia']
        numero_emergencia = request.form['numero_emergencia']

        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO empleados (nombre, cedula, cargo, valor_dia, numero_emergencia) VALUES (%s, %s, %s, %s, %s)",
                    (nombre, cedula, cargo, valor_dia, numero_emergencia))
        mysql.connection.commit()
        cur.close()
        return redirect('/empleados')
    return render_template('agregar_empleado.html')

@app.route('/asistencia', methods=['GET'])
def asistencia():
    if 'usuario' not in session:
        return redirect('/login')
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM empleados WHERE activo = TRUE")
    empleados = cur.fetchall()
    cur.execute("SELECT empleado_id FROM asistencia WHERE fecha = CURDATE()")
    registrados_hoy = [row[0] for row in cur.fetchall()]
    cur.close()
    return render_template('asistencia.html', empleados=empleados,
                         registrados_hoy=registrados_hoy, fecha=date.today())

@app.route('/reporte')
def reporte():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.cargo, a.fecha, a.estado
        FROM asistencia a
        JOIN empleados e ON a.empleado_id = e.id
        ORDER BY a.fecha, e.nombre
    """)
    datos = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    ws.append(["Nombre", "Cargo", "Fecha", "Estado"])

    for fila in datos:
        ws.append([fila[0], fila[1], str(fila[2]), fila[3]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="reporte_asistencia.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        contrasena = request.form['contrasena']
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM usuarios WHERE usuario=%s", (usuario,))
        user = cur.fetchone()
        cur.close()
        if user and check_password_hash(user[2], contrasena):
            session['usuario'] = usuario
            if user[3] == 1:  # primer_login
                return redirect('/cambiar_contrasena')
            return redirect('/dashboard')
        return render_template('login.html', error='Usuario o contraseña incorrectos')
    return render_template('login.html')

@app.route('/reporte_pago')
def reporte_pago():
    if 'usuario' not in session:
        return redirect('/login')  #listo la redirección a login 
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.cargo,
               SUM(CASE WHEN a.estado = 'presente' THEN 1 ELSE 0 END) as dias_trabajados,
               e.valor_dia,
               SUM(CASE WHEN a.estado = 'presente' THEN 1 ELSE 0 END) * e.valor_dia as total
        FROM empleados e
        LEFT JOIN asistencia a ON e.id = a.empleado_id
        GROUP BY e.id, e.nombre, e.cargo, e.valor_dia
    """)
    pagos = cur.fetchall()
    cur.close()
    return render_template('reporte_pago.html', pagos=pagos)

@app.route('/reporte_pago_excel')
def reporte_pago_excel():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT e.nombre, e.cargo,
               SUM(CASE WHEN a.estado = 'presente' THEN 1 ELSE 0 END) as dias_trabajados,
               e.valor_dia,
               SUM(CASE WHEN a.estado = 'presente' THEN 1 ELSE 0 END) * e.valor_dia as total
        FROM empleados e
        LEFT JOIN asistencia a ON e.id = a.empleado_id
        GROUP BY e.id, e.nombre, e.cargo, e.valor_dia
    """)
    pagos = cur.fetchall()
    cur.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pago Quincenal"

    # Estilos
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    left   = Alignment(horizontal="left")
    border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7")
    )

    # Encabezados
    headers = ["Nombre", "Cargo", "Dias Trabajados", "Valor Dia", "Total a Pagar"]
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border

    # Datos
    for p in pagos:
        ws.append([p[0], p[1], p[2], float(p[3]), float(p[4])])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for col, cell in enumerate(row, 1):
            cell.border    = border
            cell.alignment = left if col <= 2 else center

    # Auto-ancho
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name="pago_quincenal.xlsx",
                     as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect('/login') #listo la redireccion a login
    cur = mysql.connection.cursor()
    
    cur.execute("SELECT COUNT(*) FROM empleados WHERE activo = TRUE")
    total_empleados = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM asistencia WHERE fecha = CURDATE() AND estado = 'presente'")
    presentes_hoy = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM asistencia WHERE fecha = CURDATE() AND estado = 'ausente'")
    ausentes_hoy = cur.fetchone()[0]
    
    cur.execute("""
        SELECT SUM(CASE WHEN a.estado = 'presente' THEN e.valor_dia ELSE 0 END)
        FROM asistencia a
        JOIN empleados e ON a.empleado_id = e.id
    """)
    total_pagar = cur.fetchone()[0] or 0
    cur.close()
    
    return render_template('dashboard.html',
                         total_empleados=total_empleados,
                         presentes_hoy=presentes_hoy,
                         ausentes_hoy=ausentes_hoy,
                         total_pagar=total_pagar)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/cambiar_contrasena', methods=['GET', 'POST'])
def cambiar_contrasena():
    if 'usuario' not in session:
        return redirect('/login')
    if request.method == 'POST':
        nueva = request.form['nueva_contrasena']
        regex = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[!@#$%^&*_\-*])[A-Za-z\d!@#$%^&*_\-*]{8,}$'
        if not re.match(regex, nueva):
            return render_template('cambiar_contrasena.html',
                error='Mínimo 8 caracteres, mayúscula, minúscula, número y carácter especial (!@#$%^&*_-)')
        hash_nueva = generate_password_hash(nueva)
        cur = mysql.connection.cursor()
        cur.execute("UPDATE usuarios SET contrasena=%s, primer_login=0 WHERE usuario=%s",
                    (hash_nueva, session['usuario']))
        mysql.connection.commit()
        cur.close()
        return redirect('/dashboard')
    return render_template('cambiar_contrasena.html')

@app.route('/guardar_asistencia/<int:empleado_id>', methods=['POST'])
def guardar_asistencia(empleado_id):
    if 'usuario' not in session:
        return redirect('/login')
    estado = request.form.get('estado')
    hora_llegada = datetime.now().time() if estado == 'tarde' else None
    try:
        cur = mysql.connection.cursor()
        cur.execute(
            "INSERT INTO asistencia (empleado_id, fecha, estado, hora_llegada) VALUES (%s, CURDATE(), %s, %s)",
            (empleado_id, estado, hora_llegada)
        )
        mysql.connection.commit()
        cur.close()
    except Exception:
        mysql.connection.rollback()
    return redirect('/asistencia')

if __name__ == '__main__':
   app.run(debug=True, host='0.0.0.0')